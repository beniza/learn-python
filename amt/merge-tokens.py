'''
Module to merge a set of individual token files stored in the current directory 
to one single Spreadsheet (and a text file)
'''
import openpyxl
import glob

files = glob.glob("*.xlsx")
tok = {}

for f in files:
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    for r in range(ws.max_row):
        try:
            tok[ws["A" + str(r+1)].value].append(ws["B" + str(r+1)].value)
        except KeyError:
            tok[ws["A" + str(r+1)].value] = ws["B" + str(r+1)].value
        except AttributeError:
            pass

o = open("tokens.txt", mode='w', encoding='utf-8')
try:
    for k, v in tok.items():
        o.write(str(k) + "\t" + str(v) + "\n")
except:
    pass
o.close()

out_wb = openpyxl.Workbook()
out_ws = out_wb.active
try:
    for row, (k, v) in enumerate(tok.items()):
        out_ws["A" + str(row + 1)].value = k
        out_ws["B" + str(row + 1)].value = v
except:
    out_ws["A" + str(row + 1)].value = "###"
    out_wb["B" + str(row + 1)].value = "***"
    pass

out_wb.save("Bodo_Parjo.xlsx")
print("\n" + str(len(tok)) + " tokens are saved to the file.")
